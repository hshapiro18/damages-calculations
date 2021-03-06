#Underpayment Calculator
#Created by Ryan Abele
from flask import Flask
app = Flask(_name_)
from openpyxl import Workbook
import datetime
import re
import math
from openpyxl.compat import basestring
from openpyxl.utils.exceptions import CellCoordinatesException

@app.route("/")

def main():

	tipped=is_Tipped()
	days=number_Days()
	spread=include_Spread_Of_Hours()
	workbook_Name=get_Workbook_Name()
	if tipped==True:
		if spread==True:
			Tipped_Spread_Worksheet(days, workbook_Name)
		else:
			Tipped_Worksheet(days, workbook_Name)
	else:
		if spread==True:
			non_Tipped_Spread_Worksheet(days, workbook_Name)
		else:
			non_Tipped_Worksheet(days, workbook_Name)
	
def is_Tipped():
	print("Is the employee tipped? Enter 'Y' for yes, 'N' for no: ")
	tipped=str(input())
	tipped.upper()
	if tipped=="Y" or tipped=="y":
		return True
	else:
		return False

def number_Days():
	print("Please enter the number of days in the pay period: ")
	days=int(input())
	return days
	
def include_Spread_Of_Hours():
	print("Would you like to include Spread of Hours Rights into the spreadsheet? Enter 'Y' for yes, 'N' for no: ")
	spread=str(input())
	spread.upper()
	if spread=="Y" or spread=="y":
		return True
	else:
		return False
	
def get_Workbook_Name():
	print("Please enter the title for your new workbook: ")
	title=str(input())
	title+=".xlsx"
	return title

def Tipped_Worksheet(payperiod_Length, workbook_Name):
	
	wb=Workbook()
	ws=wb.active
	ws.title="Underpayment Calculator"
	
	employee_Title="Employee"
	ws.cell(row=1, column=1, value=employee_Title)
	wage_Title="Wage"
	ws.cell(row=1, column=2, value=wage_Title)
	min_Wage_Title="Minimum Wage"
	ws.cell(row=1, column=3, value=min_Wage_Title)
	
	day=1
	for col in range(4, payperiod_Length+4):
		ws.cell(row=1, column=col, value="Day "+str(day))
		ws.cell(row=1, column=col+payperiod_Length, value=("Day "+str(day)+" Call-In Hours"))
		ws.cell(row=2, column=col+payperiod_Length, value=("=IF(AND("+_get_column_letter(col)+"2<4, "+_get_column_letter(col)+"2>0), 4-"+_get_column_letter(col)+"2, 0)"))
		day=day+1
	
	weeks=int(math.ceil(payperiod_Length/7))
	
	for i in range(0, weeks):
		week_Num="Week "+str(i+1)
		week_Hours_Title=week_Num+" Hours"
		ws.cell(row=1, column=payperiod_Length*2+4+i, value=week_Hours_Title)
		title="=SUM("+_get_column_letter(4+7*i)+"2:"+_get_column_letter(10+7*i)+"2)"
		ws.cell(row=2, column=payperiod_Length*2+4+i, value=title)
		
		week_Overtime_Title=week_Num+" Overtime Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+weeks+i, value=week_Overtime_Title)
		next_Value="=IF("+_get_column_letter(payperiod_Length*2+4+i)+"2>40, "+_get_column_letter(payperiod_Length*2+4+i)+"2-40, 0)"
		ws.cell(row=2, column=payperiod_Length*2+5+weeks+i, value=next_Value)
		
		week_Regular_Title=week_Num+" Regular Hours"
		ws.cell(row=1, column=payperiod_Length*2+6+weeks*2+i, value=week_Regular_Title)
		ws.cell(row=2, column=payperiod_Length*2+6+weeks*2+i, value="="+_get_column_letter(payperiod_Length*2+4+i)+"2-"+_get_column_letter(payperiod_Length*2+5+weeks+i)+"2")
		
		i+=1
		
			
	total_Hours_Title="Total Hours"
	ws.cell(row=1, column=payperiod_Length*weeks+3+weeks, value=total_Hours_Title)
	overtime_Hours_Title="Total Overtime Hours"
	ws.cell(row=1, column=payperiod_Length*2+5+weeks*2, value=overtime_Hours_Title)
	regular_Hours_Title="Total Regular Hours"
	ws.cell(row=1, column=payperiod_Length*2+6+weeks*3, value=regular_Hours_Title)
	
	total_Callin_Title="Total Call-In Hours"
	ws.cell(row=1, column=payperiod_Length*2+7+weeks*3, value=total_Callin_Title)
	total_Tips_Title="Total Tips"
	ws.cell(row=1, column=payperiod_Length*2+8+weeks*3, value=total_Tips_Title)
	total_Paid_Title="Total Paid"
	ws.cell(row=1, column=payperiod_Length*2+9+weeks*3, value=total_Paid_Title)
	total_Due_Title="Total Due"
	ws.cell(row=1, column=payperiod_Length*2+10+weeks*3, value=total_Due_Title)
	total_Owed_Title="Total Owed"
	ws.cell(row=1, column=payperiod_Length*2+11+weeks*3, value=total_Owed_Title)
	
	last_Day_Column_Letter=_get_column_letter(payperiod_Length+3)
	total_Hours_Column_Letter=_get_column_letter(payperiod_Length*2+4+weeks)
	total_Overtime_Column_Letter=_get_column_letter(payperiod_Length*2+5+weeks*2)
	total_Regular_Column_Letter=_get_column_letter(payperiod_Length*2+6+weeks*3)
	
	total_Hours="=SUM(D2:"+last_Day_Column_Letter+"2)"
	overtime_Hours="=SUM("+_get_column_letter(payperiod_Length*2+5+weeks)+"2:"+_get_column_letter(payperiod_Length*2+5+weeks+(weeks-1))+"2)"
	regular_Hours="=SUM("+_get_column_letter(payperiod_Length*2+6+weeks*2)+"2:"+_get_column_letter(payperiod_Length*2+6+weeks*2+(weeks-1))+"2)"
	
	ws.cell(row=2, column=payperiod_Length*2+4+weeks, value=total_Hours)
	ws.cell(row=2, column=payperiod_Length*2+5+weeks*2, value=overtime_Hours)
	ws.cell(row=2, column=payperiod_Length*2+6+weeks*3, value=regular_Hours)
	
	total_Callin_Column_Letter=_get_column_letter(payperiod_Length*2+7+weeks*3)
	total_Tips_Column_Letter=_get_column_letter(payperiod_Length*2+8+weeks*3)
	total_Pay_Column_Letter=_get_column_letter(payperiod_Length*2+9+weeks*3)
	total_Due_Column_Letter=_get_column_letter(payperiod_Length*2+10+weeks*3)
	total_Owed_Column_Letter=_get_column_letter(payperiod_Length*2+11+weeks*3)
	
	total_Callin="=SUM("+_get_column_letter(4+payperiod_Length)+"2:"+_get_column_letter(3+payperiod_Length*2)+"2)"	
	minwage_comp="C2*"+total_Regular_Column_Letter+"2+C2*1.5*"+total_Overtime_Column_Letter+"2"
	with_tip_comp="B2*"+total_Regular_Column_Letter+"2+B2*1.5*"+total_Overtime_Column_Letter+"2+"+total_Tips_Column_Letter+"2"
	if_min_greater="C2*"+total_Regular_Column_Letter+"2+C2*1.5*"+total_Overtime_Column_Letter+"2+"+total_Callin_Column_Letter+"2*C2"
	if_min_less="B2*"+total_Regular_Column_Letter+"2+B2*1.5*"+total_Overtime_Column_Letter+"2+"+total_Tips_Column_Letter+"2+"+total_Callin_Column_Letter+"2*C2"
	total_Due="=IF("+minwage_comp+">"+with_tip_comp+", "+if_min_greater+", "+if_min_less+")"
	total_Owed="="+total_Due_Column_Letter+"2-"+total_Pay_Column_Letter+"2"
	
	ws.cell(row=2, column=payperiod_Length*2+7+weeks*3, value=total_Callin)
	ws.cell(row=2, column=payperiod_Length*2+10+weeks*3, value=total_Due)
	ws.cell(row=2, column=payperiod_Length*2+11+weeks*3, value=total_Owed)
	
	wb.save(workbook_Name)

def Tipped_Spread_Worksheet(payperiod_Length, workbook_Name):
		
	wb=Workbook()
	ws=wb.active
	ws.title="Underpayment Calculator"
	
	employee_Title="Employee"
	ws.cell(row=1, column=1, value=employee_Title)
	wage_Title="Wage"
	ws.cell(row=1, column=2, value=wage_Title)
	min_Wage_Title="Minimum Wage"
	ws.cell(row=1, column=3, value=min_Wage_Title)
	
	day=1
	for col in range(4, payperiod_Length+4):
		ws.cell(row=1, column=col, value="Day "+str(day))
		ws.cell(row=1, column=col+payperiod_Length, value=("Day "+str(day)+" Call-In Hours"))
		ws.cell(row=2, column=col+payperiod_Length, value=("=IF(AND("+_get_column_letter(col)+"2<4, "+_get_column_letter(col)+"2>0), 4-"+_get_column_letter(col)+"2, 0)"))
		day=day+1
	
	weeks=int(payperiod_Length/7)
	
	for i in range(0, weeks):
		week_Num="Week "+str(i+1)
		week_Hours_Title=week_Num+" Hours"
		ws.cell(row=1, column=payperiod_Length*2+4+i, value=week_Hours_Title)
		title="=SUM("+_get_column_letter(4+7*i)+"2:"+_get_column_letter(10+7*i)+"2)"
		ws.cell(row=2, column=payperiod_Length*2+4+i, value=title)
		
		week_Overtime_Title=week_Num+" Overtime Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+weeks+i, value=week_Overtime_Title)
		next_Value="=IF("+_get_column_letter(payperiod_Length*2+4+i)+"2>40, "+_get_column_letter(payperiod_Length*2+4+i)+"2-40, 0)"
		ws.cell(row=2, column=payperiod_Length*2+5+weeks+i, value=next_Value)
		
		week_Regular_Title=week_Num+" Regular Hours"
		ws.cell(row=1, column=payperiod_Length*2+6+weeks*2+i, value=week_Regular_Title)
		ws.cell(row=2, column=payperiod_Length*2+6+weeks*2+i, value="="+_get_column_letter(payperiod_Length*2+4+i)+"2-"+_get_column_letter(payperiod_Length*2+5+weeks+i)+"2")
		
		i+=1
		
			
	total_Hours_Title="Total Hours"
	ws.cell(row=1, column=payperiod_Length*2+4+weeks, value=total_Hours_Title)
	overtime_Hours_Title="Total Overtime Hours"
	ws.cell(row=1, column=payperiod_Length*2+5+weeks*2, value=overtime_Hours_Title)
	regular_Hours_Title="Total Regular Hours"
	ws.cell(row=1, column=payperiod_Length*2+6+weeks*3, value=regular_Hours_Title)
	
	total_Spread_Title="Total Days with SOH"
	ws.cell(row=1, column=payperiod_Length*2+7+weeks*3, value=total_Spread_Title)
	total_Callin_Title="Total Call-In Hours"
	ws.cell(row=1, column=payperiod_Length*2+8+weeks*3, value=total_Callin_Title)
	total_Tips_Title="Total Tips"
	ws.cell(row=1, column=payperiod_Length*2+9+weeks*3, value=total_Tips_Title)
	total_Paid_Title="Total Paid"
	ws.cell(row=1, column=payperiod_Length*2+10+weeks*3, value=total_Paid_Title)
	total_Due_Title="Total Due"
	ws.cell(row=1, column=payperiod_Length*2+11+weeks*3, value=total_Due_Title)
	total_Owed_Title="Total Owed"
	ws.cell(row=1, column=payperiod_Length*2+12+weeks*3, value=total_Owed_Title)
	
	last_Day_Column_Letter=_get_column_letter(payperiod_Length+3)
	total_Hours_Column_Letter=_get_column_letter(payperiod_Length*2+4+weeks)
	total_Overtime_Column_Letter=_get_column_letter(payperiod_Length*2+5+weeks*2)
	total_Regular_Column_Letter=_get_column_letter(payperiod_Length*2+6+weeks*3)
	
	total_Hours="=SUM(D2:"+last_Day_Column_Letter+"2)"
	overtime_Hours="=SUM("+_get_column_letter(payperiod_Length*2+5+weeks)+"2:"+_get_column_letter(payperiod_Length*2+5+weeks+(weeks-1))+"2)"
	regular_Hours="=SUM("+_get_column_letter(payperiod_Length*2+6+weeks*2)+"2:"+_get_column_letter(payperiod_Length*2+6+weeks*2+(weeks-1))+"2)"
	
	ws.cell(row=2, column=payperiod_Length*2+4+weeks, value=total_Hours)
	ws.cell(row=2, column=payperiod_Length*2+5+weeks*2, value=overtime_Hours)
	ws.cell(row=2, column=payperiod_Length*2+6+weeks*3, value=regular_Hours)
	
	total_Spread_Column_Letter=_get_column_letter(payperiod_Length*2+7+weeks*3)
	total_Callin_Column_Letter=_get_column_letter(payperiod_Length*2+8+weeks*3)
	total_Tips_Column_Letter=_get_column_letter(payperiod_Length*2+9+weeks*3)
	total_Pay_Column_Letter=_get_column_letter(payperiod_Length*2+10+weeks*3)
	total_Due_Column_Letter=_get_column_letter(payperiod_Length*2+11+weeks*3)
	total_Owed_Column_Letter=_get_column_letter(payperiod_Length*2+12+weeks*3)
	
	total_Spread='=COUNTIF('+_get_column_letter(4)+'2:'+_get_column_letter(3+payperiod_Length)+'2, ">10")'
	total_Callin="=SUM("+_get_column_letter(4+payperiod_Length)+"2:"+_get_column_letter(3+payperiod_Length*2)+"2)"	
	minwage_comp="C2*"+total_Regular_Column_Letter+"2+C2*1.5*"+total_Overtime_Column_Letter+"2"
	with_tip_comp="B2*"+total_Regular_Column_Letter+"2+B2*1.5*"+total_Overtime_Column_Letter+"2+"+total_Tips_Column_Letter+"2"
	if_min_greater="C2*"+total_Regular_Column_Letter+"2+C2*1.5*"+total_Overtime_Column_Letter+"2+"+total_Callin_Column_Letter+"2*C2+"+total_Spread_Column_Letter+"2*C2"
	if_min_less="B2*"+total_Regular_Column_Letter+"2+B2*1.5*"+total_Overtime_Column_Letter+"2+"+total_Tips_Column_Letter+"2+"+total_Callin_Column_Letter+"2*C2+"+total_Spread_Column_Letter+"2*C2"
	total_Due="=IF("+minwage_comp+">"+with_tip_comp+", "+if_min_greater+", "+if_min_less+")"
	total_Owed="="+total_Due_Column_Letter+"2-"+total_Pay_Column_Letter+"2"
	
	ws.cell(row=2, column=payperiod_Length*2+7+weeks*3, value=total_Spread)
	ws.cell(row=2, column=payperiod_Length*2+8+weeks*3, value=total_Callin)
	ws.cell(row=2, column=payperiod_Length*2+11+weeks*3, value=total_Due)
	ws.cell(row=2, column=payperiod_Length*2+12+weeks*3, value=total_Owed)
	
	wb.save(workbook_Name)

def non_Tipped_Worksheet(payperiod_Length, workbook_Name):
	
	wb=Workbook()
	ws=wb.active
	ws.title="Underpayment Calculator"
	
	employee_Title="Employee"
	ws.cell(row=1, column=1, value=employee_Title)
	wage_Title="Wage"
	ws.cell(row=1, column=2, value=wage_Title)
	minwage_Title="Minimum Wage"
	ws.cell(row=1, column=3, value=minwage_Title)
	realwage_Title="Real Wage"
	ws.cell(row=1, column=4, value=realwage_Title)
	ws.cell(row=2, column=4, value="=IF(B2>C2, B2, C2)")
	
	day=1
	for col in range(5, payperiod_Length+5):
		ws.cell(row=1, column=col, value="Day "+str(day))
		ws.cell(row=1, column=col+payperiod_Length, value=("Day "+str(day)+" Call-In Hours"))
		ws.cell(row=2, column=col+payperiod_Length, value=("=IF(AND("+_get_column_letter(col)+"2<4, "+_get_column_letter(col)+"2>0), 4-"+_get_column_letter(col)+"2, 0)"))
		day=day+1
		
	weeks=int(payperiod_Length/7)
	
	for i in range(0, weeks):
		week_Num="Week "+str(i+1)
		week_Hours_Title=week_Num+" Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+i, value=week_Hours_Title)
		title="=SUM("+_get_column_letter(5+7*i)+"2:"+_get_column_letter(11+7*i)+"2)"
		ws.cell(row=2, column=payperiod_Length*2+5+i, value=title)
		
		week_Overtime_Title=week_Num+" Overtime Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+weeks+i, value=week_Overtime_Title)
		next_Value="=IF("+_get_column_letter(payperiod_Length*2+5+i)+"2>40, "+_get_column_letter(payperiod_Length*2+5+i)+"2-40, 0)"
		ws.cell(row=2, column=payperiod_Length*2+5+weeks+i, value=next_Value)
		
		week_Regular_Title=week_Num+" Regular Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+weeks*2+i, value=week_Regular_Title)
		ws.cell(row=2, column=payperiod_Length*2+5+weeks*2+i, value="="+_get_column_letter(payperiod_Length*2+5+i)+"2-"+_get_column_letter(payperiod_Length*2+5+weeks+i)+"2")
		
		i+=1
	
	total_Hours_Title="Total Hours"
	ws.cell(row=1, column=payperiod_Length*2+5+weeks*3, value=total_Hours_Title)
	
	overtime_Hours_Title="Total Overtime Hours"
	ws.cell(row=1, column=payperiod_Length*2+6+weeks*3, value=overtime_Hours_Title)
	regular_Hours_Title="Total Regular Hours"
	ws.cell(row=1, column=payperiod_Length*2+7+weeks*3, value=regular_Hours_Title)
	
	total_Callin_Title="Total Call in Hours"
	ws.cell(row=1, column=payperiod_Length*2+8+weeks*3, value=total_Callin_Title)
	total_Paid_Title="Total Paid"
	ws.cell(row=1, column=payperiod_Length*2+9+weeks*3, value=total_Paid_Title)
	total_Due_Title="Total Due"
	ws.cell(row=1, column=payperiod_Length*2+10+weeks*3, value=total_Due_Title)
	total_Owed_Title="Total Owed"
	ws.cell(row=1, column=payperiod_Length*2+11+weeks*3, value=total_Owed_Title)
	
	last_Day_Column_Letter=_get_column_letter(payperiod_Length+4)
	total_Hours_Column_Letter=_get_column_letter(payperiod_Length*2+5+weeks*3)
	total_Overtime_Column_Letter=_get_column_letter(payperiod_Length*2+6+weeks*3)
	total_Regular_Column_Letter=_get_column_letter(payperiod_Length*2+7+weeks*3)
	
	total_Hours="=SUM(E2:"+last_Day_Column_Letter+"2)"
	overtime_Hours="=SUM("+_get_column_letter(payperiod_Length*2+5+weeks)+"2:"+_get_column_letter(payperiod_Length*2+5+weeks+(weeks-1))+"2)"
	regular_Hours="=SUM("+_get_column_letter(payperiod_Length*2+5+weeks*2)+"2:"+_get_column_letter(payperiod_Length*2+5+(weeks-1)+weeks*2)+"2)"
	
	ws.cell(row=2, column=payperiod_Length*2+5+weeks*3, value=total_Hours)
	ws.cell(row=2, column=payperiod_Length*2+6+weeks*3, value=overtime_Hours)
	ws.cell(row=2, column=payperiod_Length*2+7+weeks*3, value=regular_Hours)
	
	total_Callin_Column_Letter=_get_column_letter(payperiod_Length*2+8+weeks*3)
	total_Pay_Column_Letter=_get_column_letter(payperiod_Length*2+9+weeks*3)
	total_Due_Column_Letter=_get_column_letter(payperiod_Length*2+10+weeks*3)
	total_Owed_Column_Letter=_get_column_letter(payperiod_Length*2+11+weeks*3)
	
	total_Callin="=SUM("+_get_column_letter(5+payperiod_Length)+"2:"+_get_column_letter(4+payperiod_Length*2)+"2)"		
	total_Due="=(D2*"+total_Regular_Column_Letter+"2)+(D2*1.5*"+total_Overtime_Column_Letter+"2)+("+total_Callin_Column_Letter+"2*C2)"
	total_Owed="="+total_Due_Column_Letter+"2-"+total_Pay_Column_Letter+"2"
	
	ws.cell(row=2, column=payperiod_Length*2+8+weeks*3, value=total_Callin)
	ws.cell(row=2, column=payperiod_Length*2+10+weeks*3, value=total_Due)
	ws.cell(row=2, column=payperiod_Length*2+11+weeks*3, value=total_Owed)
	
	wb.save(workbook_Name)
	
def non_Tipped_Spread_Worksheet(payperiod_Length, workbook_Name):
	
	wb=Workbook()
	ws=wb.active
	ws.title="Underpayment Calculator"
	
	employee_Title="Employee"
	ws.cell(row=1, column=1, value=employee_Title)
	wage_Title="Wage"
	ws.cell(row=1, column=2, value=wage_Title)
	minwage_Title="Minimum Wage"
	ws.cell(row=1, column=3, value=minwage_Title)
	realwage_Title="Real Wage"
	ws.cell(row=1, column=4, value=realwage_Title)
	ws.cell(row=2, column=4, value="=IF(B2>C2, B2, C2)")
	
	day=1
	for col in range(5, payperiod_Length+5):
		ws.cell(row=1, column=col, value="Day "+str(day))
		ws.cell(row=1, column=col+payperiod_Length, value=("Day "+str(day)+" Call-In Hours"))
		ws.cell(row=2, column=col+payperiod_Length, value=("=IF(AND("+_get_column_letter(col)+"2<4, "+_get_column_letter(col)+"2>0), 4-"+_get_column_letter(col)+"2, 0)"))
		day=day+1
		
	weeks=int(payperiod_Length/7)
	
	for i in range(0, weeks):
		week_Num="Week "+str(i+1)
		week_Hours_Title=week_Num+" Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+i, value=week_Hours_Title)
		title="=SUM("+_get_column_letter(5+7*i)+"2:"+_get_column_letter(11+7*i)+"2)"
		ws.cell(row=2, column=payperiod_Length*2+5+i, value=title)
		
		week_Overtime_Title=week_Num+" Overtime Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+weeks+i, value=week_Overtime_Title)
		next_Value="=IF("+_get_column_letter(payperiod_Length*2+5+i)+"2>40, "+_get_column_letter(payperiod_Length*2+5+i)+"2-40, 0)"
		ws.cell(row=2, column=payperiod_Length*2+5+weeks+i, value=next_Value)
		
		week_Regular_Title=week_Num+" Regular Hours"
		ws.cell(row=1, column=payperiod_Length*2+5+weeks*2+i, value=week_Regular_Title)
		ws.cell(row=2, column=payperiod_Length*2+5+weeks*2+i, value="="+_get_column_letter(payperiod_Length*2+5+i)+"2-"+_get_column_letter(payperiod_Length*2+5+weeks+i)+"2")
		
		i+=1
	
	total_Hours_Title="Total Hours"
	ws.cell(row=1, column=payperiod_Length*2+5+weeks*3, value=total_Hours_Title)
	
	overtime_Hours_Title="Total Overtime Hours"
	ws.cell(row=1, column=payperiod_Length*2+6+weeks*3, value=overtime_Hours_Title)
	regular_Hours_Title="Total Regular Hours"
	ws.cell(row=1, column=payperiod_Length*2+7+weeks*3, value=regular_Hours_Title)
	
	total_SOH_Title="Total Days with SOH"
	ws.cell(row=1, column=payperiod_Length*2+8+weeks*3, value=total_SOH_Title)
	total_Callin_Title="Total Call in Hours"
	ws.cell(row=1, column=payperiod_Length*2+9+weeks*3, value=total_Callin_Title)
	total_Paid_Title="Total Paid"
	ws.cell(row=1, column=payperiod_Length*2+10+weeks*3, value=total_Paid_Title)
	total_Due_Title="Total Due"
	ws.cell(row=1, column=payperiod_Length*2+11+weeks*3, value=total_Due_Title)
	total_Owed_Title="Total Owed"
	ws.cell(row=1, column=payperiod_Length*2+12+weeks*3, value=total_Owed_Title)
	
	last_Day_Column_Letter=_get_column_letter(payperiod_Length+4)
	total_Hours_Column_Letter=_get_column_letter(payperiod_Length*2+5+weeks*3)
	total_Overtime_Column_Letter=_get_column_letter(payperiod_Length*2+6+weeks*3)
	total_Regular_Column_Letter=_get_column_letter(payperiod_Length*2+7+weeks*3)
	
	total_Hours="=SUM(E2:"+last_Day_Column_Letter+"2)"
	overtime_Hours="=SUM("+_get_column_letter(payperiod_Length*2+5+weeks)+"2:"+_get_column_letter(payperiod_Length*2+5+weeks+(weeks-1))+"2)"
	regular_Hours="=SUM("+_get_column_letter(payperiod_Length*2+5+weeks*2)+"2:"+_get_column_letter(payperiod_Length*2+5+(weeks-1)+weeks*2)+"2)"
	
	ws.cell(row=2, column=payperiod_Length*2+5+weeks*3, value=total_Hours)
	ws.cell(row=2, column=payperiod_Length*2+6+weeks*3, value=overtime_Hours)
	ws.cell(row=2, column=payperiod_Length*2+7+weeks*3, value=regular_Hours)
	
	total_SOH_Column_Letter=_get_column_letter(payperiod_Length*2+8+weeks*3)
	total_Callin_Column_Letter=_get_column_letter(payperiod_Length*2+9+weeks*3)
	total_Pay_Column_Letter=_get_column_letter(payperiod_Length*2+10+weeks*3)
	total_Due_Column_Letter=_get_column_letter(payperiod_Length*2+11+weeks*3)
	total_Owed_Column_Letter=_get_column_letter(payperiod_Length*2+12+weeks*3)
	
	total_Spread='=COUNTIF('+_get_column_letter(5)+'2:'+_get_column_letter(4+payperiod_Length)+'2, ">10")'
	total_Callin="=SUM("+_get_column_letter(5+payperiod_Length)+"2:"+_get_column_letter(4+payperiod_Length*2)+"2)"		
	total_Due="=(D2*"+total_Regular_Column_Letter+"2)+(D2*1.5*"+total_Overtime_Column_Letter+"2)+("+total_Callin_Column_Letter+"2*C2)+("+total_SOH_Column_Letter+"2*C2)"
	total_Owed="="+total_Due_Column_Letter+"2-"+total_Pay_Column_Letter+"2"
	
	ws.cell(row=2, column=payperiod_Length*2+8+weeks*3, value=total_Spread)
	ws.cell(row=2, column=payperiod_Length*2+9+weeks*3, value=total_Callin)
	ws.cell(row=2, column=payperiod_Length*2+11+weeks*3, value=total_Due)
	ws.cell(row=2, column=payperiod_Length*2+12+weeks*3, value=total_Owed)
	
	wb.save(workbook_Name)
	
def _get_column_letter(col_idx):

    # these indicies corrospond to A -> ZZZ and include all allowed
    # columns
    if not 1 <= col_idx <= 18278:
        raise ValueError("Invalid column index {0}".format(col_idx))
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx, 26)
        # check for exact division and borrow if needed
        if remainder == 0:
            remainder = 26
            col_idx -= 1
        letters.append(chr(remainder+64))
    return ''.join(reversed(letters))
    
_COL_STRING_CACHE = {}
_STRING_COL_CACHE = {}
for i in range(1, 18279):
    col = _get_column_letter(i)
    _STRING_COL_CACHE[i] = col
    _COL_STRING_CACHE[col] = i

if _name_ == "_main_":
	app.run()
		
main()


	
	
	

	
	
	
	
	
	